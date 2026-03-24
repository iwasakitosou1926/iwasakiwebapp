import os
import re
import json
import time
import math
import random
import tempfile
import shutil
import platform
import sys
import zipfile
import subprocess
from collections import defaultdict
from typing import Dict, List, Optional, Tuple
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from notion_client import Client
from notion_client.errors import APIResponseError, RequestTimeoutError

# Tkinter for optional user-choice dialogs (used when multiple part candidates exist)
try:
    import tkinter as tk
    from tkinter import messagebox
except Exception:
    tk = None
    messagebox = None


from layout_themes import (
    THEMES,
    CURRENT_THEME,
    APP_QTY_COLOR,
    APP_QTY_REQUIRE_BOLD,
    APP_QTY_ALLOW_NEIGHBORS,
    APP_SKIP_EMPTY_QTY,
)

# =====================================================
# LOAD .env AND NOTION SETTINGS
# =====================================================

from dotenv import load_dotenv


def _find_dotenv_path() -> str | None:
    """Find a stable .env path across dev + frozen EXE."""
    candidates = []

    # 1) OUTER install folder (preferred for Windows EXE)
    try:
        candidates.append(os.path.join(get_install_dir(), ".env"))
    except Exception:
        pass

    # 2) Current working directory (launcher.py chdir's to OUTER)
    try:
        candidates.append(os.path.join(os.getcwd(), ".env"))
    except Exception:
        pass

    # 3) This file's directory (dev runs)
    try:
        candidates.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))
    except Exception:
        pass

    for p in candidates:
        try:
            if p and os.path.isfile(p):
                return p
        except Exception:
            continue
    return None


_env_path = _find_dotenv_path()
if _env_path:
    load_dotenv(dotenv_path=_env_path, override=True)
else:
    # Fallback: behave like before (may rely on CWD)
    load_dotenv(override=True)


NOTION_TOKEN = os.getenv("NOTION_TOKEN")
CALENDAR_DATABASE_ID = os.getenv("CALENDAR_DATABASE_ID")
PARTS_DATABASE_ID = os.getenv("PARTS_DATABASE_ID")

try:
    if _env_path:
        log(f"[ENV] loaded .env = {_env_path}")
    else:
        log("[ENV] .env not found (using default dotenv search)")
except Exception:
    pass

def verify_notion_config():
    """Ensure all required Notion environment variables are present."""
    if not NOTION_TOKEN:
        raise RuntimeError(
            "NOTION_TOKEN is missing.\n"
            "Create a .env file and set NOTION_TOKEN=your_integration_token"
        )
    # Safety check: token must be pure ASCII (no full-width / hidden chars)
    try:
        NOTION_TOKEN.encode("ascii")
    except UnicodeEncodeError:
        raise RuntimeError(
            "NOTION_TOKEN contains non-ASCII characters.\n"
            "Please re-copy it from Notion and paste into .env using only standard characters."
        )

    if not CALENDAR_DATABASE_ID or not PARTS_DATABASE_ID:
        raise RuntimeError(
            "CALENDAR_DATABASE_ID or PARTS_DATABASE_ID is missing.\n"
            "Please set both in your .env file."
        )

# Notion client: enforce timeouts so Windows doesn't appear to freeze on network stalls.
# notion-client supports `timeout_ms` (default 60_000). We keep it modest and retry via _with_backoff.
try:
    import logging
    _notion_log_level = logging.DEBUG if os.getenv("NOTION_DEBUG", "").strip() == "1" else logging.WARNING
except Exception:
    _notion_log_level = None

client_kwargs = {"auth": NOTION_TOKEN, "timeout_ms": int(os.getenv("NOTION_TIMEOUT_MS", "30000") or "30000")}
if _notion_log_level is not None:
    client_kwargs["log_level"] = _notion_log_level

notion = Client(**client_kwargs)

# =====================================================
# LOGGING HOOK (GUI can overwrite this)
# =====================================================

_LogHandlerType = callable

def _default_log_handler(msg: str):
    # Fallback: just print to console
    print(msg)

_log_handler = _default_log_handler


def set_log_handler(handler):
    """
    Main app can call: logic.set_log_handler(app_log_function)
    so that all backend logs go into the Tk console.
    """
    global _log_handler
    _log_handler = handler or _default_log_handler


def log(text: str):
    _log_handler(str(text))

def _now_ms() -> int:
    try:
        return int(time.time() * 1000)
    except Exception:
        return 0


# =====================================================
# PATHS / CACHE
# =====================================================

def _get_resource_base_dir() -> str:
    if getattr(sys, "frozen", False):
        if hasattr(sys, "_MEIPASS"):
            return sys._MEIPASS
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def _get_user_writable_dir() -> str:
    win_appdata = os.getenv("LOCALAPPDATA") or os.getenv("APPDATA")
    if win_appdata:
        base = os.path.join(win_appdata, "NotionSyncApp")
    else:
        base = os.path.join(os.path.expanduser("~"), "NotionSyncApp")
    os.makedirs(base, exist_ok=True)
    return base


RESOURCE_BASE_DIR = _get_resource_base_dir()

USER_BASE_DIR = _get_user_writable_dir()

# =====================================================
# USER SETTINGS (persistent across runs)
# =====================================================

SETTINGS_FILE = os.path.join(USER_BASE_DIR, "settings.json")


def _load_user_settings() -> dict:
    try:
        if os.path.isfile(SETTINGS_FILE):
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                return json.load(f) or {}
    except Exception:
        pass
    return {}


def _save_user_settings(d: dict) -> None:
    try:
        os.makedirs(USER_BASE_DIR, exist_ok=True)
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(d or {}, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def get_user_selected_excel_folder() -> str | None:
    """Return user-selected Excel folder if previously saved and still exists."""
    try:
        d = _load_user_settings()
        p = str(d.get("excel_folder", "") or "").strip()
        if p and os.path.exists(p):
            return p
    except Exception:
        pass
    return None


def set_user_selected_excel_folder(path: str) -> bool:
    """Persist the chosen Excel folder and update runtime folder path."""
    try:
        p = str(path or "").strip()
        if not p or not os.path.isdir(p):
            return False
        d = _load_user_settings()
        d["excel_folder"] = p
        _save_user_settings(d)
        # also update runtime global
        global FOLDER_PATH
        FOLDER_PATH = p
        try:
            log(f"[PATH] User selected Excel folder = {p}")
        except Exception:
            pass
        return True
    except Exception:
        return False


# =====================================================
# LAN UPDATER (manifest.json + app.zip on file share)
# =====================================================

def get_install_dir() -> str:
    """Return the OUTER install folder.

    Expected on Windows (PyInstaller --onedir):
      OUTER\\NotionSyncApp\\NotionSyncApp.exe
      OUTER\\app\\...
      OUTER\\version.txt

    In dev (non-frozen): returns this file's directory.
    """
    if getattr(sys, "frozen", False):
        exe_dir = os.path.dirname(sys.executable)          # ...\\OUTER\\NotionSyncApp
        outer_dir = os.path.dirname(exe_dir)               # ...\\OUTER

        # Prefer OUTER if it looks like an install folder
        if os.path.isdir(os.path.join(outer_dir, "app")):
            return outer_dir
        if os.path.isfile(os.path.join(outer_dir, "version.txt")):
            return outer_dir

        # Fallback to exe_dir if OUTER doesn't exist / not expected layout
        return exe_dir

    # dev run (e.g., Mac / Python)
    return os.path.dirname(os.path.abspath(__file__))


def get_update_manifest_path() -> str:
    """
    Manifest location for LAN updates.

    Recommended: set env var UPDATE_MANIFEST_PATH to a UNC path, e.g.
      \\SERVER\\apps\\NotionSyncApp\\latest\\manifest.json

    Falls back to the same example path if env is not set.
    """
    envp = os.getenv("UPDATE_MANIFEST_PATH")
    if envp:
        return envp
    return r"\\\\SERVER\\apps\\NotionSyncApp\\latest\\manifest.json"


def read_local_version_from_install() -> str:
    """Reads version.txt from the OUTER install folder; returns '0.0.0' if missing."""
    try:
        p = os.path.join(get_install_dir(), "version.txt")
        if os.path.isfile(p):
            with open(p, "r", encoding="utf-8") as f:
                return f.read().strip() or "0.0.0"
    except Exception:
        pass
    return "0.0.0"


def write_local_version_to_install(version: str) -> None:
    """Write version.txt into the OUTER install folder."""
    try:
        p = os.path.join(get_install_dir(), "version.txt")
        with open(p, "w", encoding="utf-8") as f:
            f.write(str(version).strip())
    except Exception:
        # non-fatal
        pass


def _version_key(v: str) -> list:
    """Best-effort version compare.

    Supports forms like:
      2026.1.5
      2026.1.5v2
      2026.1.5-v2
      2026.1.5.2

    Strategy:
    - Split on dot/dash/underscore.
    - From each token, extract all digit groups and append them.
      e.g. '5v2' -> [5, 2]
    - If a token has no digits, append 0.
    """
    v = (v or "").strip()
    if not v:
        return [0]

    parts: list[int] = []
    tokens = re.split(r"[\.\-_]+", v)

    for tok in tokens:
        if tok is None:
            continue
        tok = str(tok).strip()
        if not tok:
            continue

        nums = re.findall(r"\d+", tok)
        if nums:
            for n in nums:
                try:
                    parts.append(int(n))
                except Exception:
                    parts.append(0)
        else:
            parts.append(0)

    return parts if parts else [0]


def is_newer_version(remote: str, local: str) -> bool:
    return _version_key(remote) > _version_key(local)


def fetch_update_manifest() -> dict:
    """Reads manifest.json from the LAN path and returns the parsed dict."""
    mp = get_update_manifest_path()
    with open(mp, "r", encoding="utf-8") as f:
        return json.load(f)


def check_update_available(current_app_version: str) -> dict:
    """
    Returns a dict:
      {
        'available': bool,
        'current': str,
        'remote': str,
        'notes': str,
        'zip_path': str,
        'manifest_path': str,
      }
    """
    manifest_path = get_update_manifest_path()
    try:
        log(f"[Update] manifest_path = {manifest_path}")
        log(f"[Update] current(APP_VERSION) = {current_app_version}")
    except Exception:
        pass
    info = {
        "available": False,
        "current": (current_app_version or "0.0.0").strip(),
        "remote": "",
        "notes": "",
        "zip_path": "",
        "manifest_path": manifest_path,
    }

    try:
        m = fetch_update_manifest()
        remote_ver = str(m.get("version", "")).strip()
        notes = str(m.get("notes", "")).strip()
        zip_name = str(m.get("zip_name", "app.zip")).strip() or "app.zip"

        # zip is alongside manifest
        base = os.path.dirname(manifest_path)
        zip_path = os.path.join(base, zip_name)

        info["remote"] = remote_ver
        info["notes"] = notes
        info["zip_path"] = zip_path

        try:
            log(f"[Update] remote(manifest) = {remote_ver}")
            log(f"[Update] zip_path = {zip_path}")
            log(f"[Update] compare: remote_key={_version_key(remote_ver)} local_key={_version_key(info['current'])}")
        except Exception:
            pass

        # Compare remote vs *APP_VERSION* (not version.txt)
        # Policy: prompt whenever the manifest version is DIFFERENT from current.
        # This allows version labels like '2026.1.5v2' to trigger updates without requiring numeric ordering.
        if remote_ver and remote_ver.strip() != info["current"].strip():
            info["available"] = True
            try:
                log("[Update] Update available (remote version differs from current).")
            except Exception:
                pass
        else:
            try:
                log("[Update] No updates (remote version matches current).")
            except Exception:
                pass

    except Exception as e:
        log(f"[Update] manifest read failed: {e}")

    return info


def _make_updater_bat(tmp_dir: str, exe_path: str, install_dir: str, extracted_dir: str, new_version: str) -> str:
    """Create a temporary updater .bat that replaces install_dir\\app and restarts the app."""
    bat_path = os.path.join(tmp_dir, "apply_update.bat")

    # Use robocopy to mirror folder (handles lots of files well).
    # We replace only the 'app' folder.
    app_src = os.path.join(extracted_dir, "app")
    app_dst = os.path.join(install_dir, "app")

    version_txt = os.path.join(install_dir, "version.txt")

    lines = [
        "@echo off",
        "setlocal",
        "timeout /t 2 /nobreak >nul",
        f"if exist \"{app_dst}\" (rmdir /s /q \"{app_dst}\")",
        f"mkdir \"{app_dst}\"",
        f"robocopy \"{app_src}\" \"{app_dst}\" /MIR /NFL /NDL /NJH /NJS /NC /NS >nul",
        f"echo {new_version} > \"{version_txt}\"",
        f"start \"\" \"{exe_path}\"",
        "endlocal",
    ]

    with open(bat_path, "w", encoding="utf-8") as f:
        f.write("\r\n".join(lines))

    return bat_path


def apply_lan_update(zip_path: str, new_version: str) -> tuple[bool, str]:
    """
    Downloads/opens the LAN zip, extracts it to a temp folder,
    launches a .bat to replace install_dir\\app, then returns.

    Returns: (ok, message)
    """
    try:
        # Only supported on Windows for your team laptops.
        if platform.system().lower() != "windows":
            return False, "Windows only updater (LAN)"

        install_dir = get_install_dir()
        exe_path = sys.executable if getattr(sys, "frozen", False) else sys.executable

        if not os.path.isfile(zip_path):
            return False, f"Update zip not found: {zip_path}"

        tmp_dir = tempfile.mkdtemp(prefix="NotionSyncAppUpdate_")
        local_zip = os.path.join(tmp_dir, "app.zip")

        # Copy from UNC to local temp
        shutil.copy2(zip_path, local_zip)

        extracted_dir = os.path.join(tmp_dir, "extract")
        os.makedirs(extracted_dir, exist_ok=True)
        with zipfile.ZipFile(local_zip, "r") as z:
            z.extractall(extracted_dir)

        # Validate structure
        if not os.path.isdir(os.path.join(extracted_dir, "app")):
            return False, "Invalid update zip: expected top-level 'app' folder."

        bat = _make_updater_bat(tmp_dir, exe_path, install_dir, extracted_dir, new_version)

        # Launch updater and exit current process ASAP
        subprocess.Popen(["cmd", "/c", bat], cwd=install_dir, creationflags=0x00000008)
        return True, "Updater launched"

    except Exception as e:
        return False, str(e)


def app_base_dir() -> str:
    return RESOURCE_BASE_DIR


BASE_DIR = app_base_dir()


def resolve_folder_path() -> str:
    """
    Look for the shared paint Excel folder (mac/Windows/UNC).
    Fallback to current working directory.
    """
    # 0) User-selected folder (highest priority)
    user_sel = get_user_selected_excel_folder()
    if user_sel:
        log(f"[PATH] using user-selected folder = {user_sel}")
        return user_sel

    env_path = os.getenv("NOTION_APP_FOLDER")
    if env_path and os.path.exists(env_path):
        log(f"[PATH] using env NOTION_APP_FOLDER = {env_path}")
        return env_path

    # Prefer UNC on Windows because mapped drives (Z:) may not exist for EXE processes.
    if platform.system().lower() == "windows":
        candidates = [
            r"\\\\SERVERNAME\\ﾃﾞｰﾀﾍﾞｰｽ\\受注センター\\フレクシェ\\塗装",  # windows UNC (most reliable)
            r"Z:\\受注センター\\フレクシェ\\塗装",                             # windows mapped drive (may not exist)
        ]
    else:
        candidates = [
            "/Volumes/ﾃﾞｰﾀﾍﾞｰｽ/受注ｾﾝﾀｰ/フレクシェ/塗装",  # mac
        ]
    for p in candidates:
        if os.path.exists(p):
            log(f"[PATH] using detected folder = {p}")
            return p

    cwd = os.getcwd()
    log(f"[PATH] fallback to CWD = {cwd}")
    log("[PATH] Hint: set NOTION_APP_FOLDER in .env to a UNC path if auto-detect fails.")
    return cwd


FOLDER_PATH = resolve_folder_path()
pattern = re.compile(r"^\d+_塗装(_[A-Za-z0-9\u3040-\u30FF\u4E00-\u9FFF_-]+)*\.xlsx$")

# cache
CACHE_DIR = os.path.join(USER_BASE_DIR, "cache")
PARTS_CACHE_FILE = os.path.join(CACHE_DIR, "parts_cache.json")
NESTED_DB_CACHE_FILE = os.path.join(CACHE_DIR, "nested_db_cache.json")
CACHE_TTL_SECONDS = 60 * 60 * 12


def _ensure_cache_dir():
    if not os.path.isdir(CACHE_DIR):
        os.makedirs(CACHE_DIR, exist_ok=True)


def _is_cache_fresh(path: str, ttl_seconds: int) -> bool:
    return os.path.isfile(path) and (time.time() - os.path.getmtime(path) < ttl_seconds)


# =====================================================
# BASIC HELPERS / BACKOFF
# =====================================================

def _with_backoff(func, *args, **kwargs):
    last_err: Exception | None = None

    for attempt in range(5):
        try:
            return func(*args, **kwargs)

        except RequestTimeoutError as e:
            last_err = e
            wait = min(2 ** attempt, 8)
            log(f"[Notion] RequestTimeout; retry in {wait}s…")
            time.sleep(wait)
            continue

        except APIResponseError as e:
            last_err = e
            if e.status in (429, 500, 502, 503, 504):
                wait = min(2 ** attempt, 8)
                log(f"Notion {e.status}; retry in {wait}s…")
                time.sleep(wait)
                continue
            # Non-retriable API error
            raise

        except Exception as e:
            last_err = e
            wait = min((2 ** attempt) + random.random(), 8)
            log(f"[Notion] Error: {type(e).__name__}; retry in {wait:.1f}s…")
            time.sleep(wait)

    # After retries, raise the last error so the UI can show it instead of hanging.
    if last_err is not None:
        raise last_err
    raise RuntimeError("Unknown error in _with_backoff")

# =====================================================
# NOTION DB QUERY COMPAT (databases.query vs data_sources.query)
# =====================================================

# Cache: database_id -> data_source_id
_DATA_SOURCE_ID_CACHE: Dict[str, str] = {}


def _get_data_source_id_for_database(database_id: str) -> str:
    """Resolve (and cache) the first data_source_id for a Notion database.

    Newer notion_client versions query databases via data_sources.query(data_source_id, ...).
    """
    if database_id in _DATA_SOURCE_ID_CACHE:
        return _DATA_SOURCE_ID_CACHE[database_id]

    db = _with_backoff(notion.databases.retrieve, database_id=database_id)
    ds_list = db.get("data_sources") or []

    ds_id = None
    if ds_list and isinstance(ds_list, list) and isinstance(ds_list[0], dict):
        ds_id = ds_list[0].get("id")

    if not ds_id:
        raise RuntimeError(
            "This notion_client build requires data_sources.query, but databases.retrieve() "
            "returned no data_sources. Check SDK version / permissions."
        )

    _DATA_SOURCE_ID_CACHE[database_id] = ds_id
    return ds_id


def notion_query_database(database_id: str, **kwargs) -> dict:
    """Query a Notion database in a version-compatible way.

    - Older SDK: notion.databases.query(database_id=..., ...)
    - Newer SDK: notion.data_sources.query(data_source_id, ...)

    Returns the raw response dict.
    """
    verify_notion_config()

    # Old SDK path (mac in your environment)
    if hasattr(notion.databases, "query"):
        return _with_backoff(notion.databases.query, database_id=database_id, **kwargs)

    # New SDK path (windows in your environment)
    if not hasattr(notion, "data_sources") or not hasattr(notion.data_sources, "query"):
        raise RuntimeError(
            "Unsupported notion_client: missing databases.query and data_sources.query. "
            "Please reinstall notion-client."
        )

    ds_id = _get_data_source_id_for_database(database_id)
    # data_sources.query signature is query(data_source_id, **kwargs)
    return _with_backoff(notion.data_sources.query, ds_id, **kwargs)

# =====================================================
# NORMALIZERS
# =====================================================

def normalize_part_key(name: str) -> str:
    if not name:
        return ""
    s = str(name)
    s = s.replace("　", " ")
    s = re.sub(r"\s+", " ", s)
    s = s.replace("－", "-").replace("―", "-").replace("ー", "-").replace("‐", "-")
    s = re.sub(r"\s*-\s*", "-", s)
    return s.strip()


def normalize_color_key(name: str) -> str:
    if not name:
        return ""
    s = str(name)
    s = s.replace("　", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalize_threef_black(s: str) -> str:
    """Return '3F黒' if s matches any 3分艶ブラック / 3F黒 variant (half/full width, spacing)."""
    if not s:
        return ""
    s0 = str(s)
    s0 = s0.replace("　", " ")
    s0 = re.sub(r"\s+", "", s0)
    trans = str.maketrans({"３": "3", "Ｆ": "F", "ｆ": "f"})
    s0 = s0.translate(trans)
    s0 = s0.replace("ブラック", "ﾌﾞﾗｯｸ").replace("ツヤ", "艶")
    if re.search(r"3[fF]黒", s0):
        return "3F黒"
    if re.search(r"3分艶(ﾌﾞﾗｯｸ|黒)", s0):
        return "3F黒"
    return s


def strip_trailing_color_suffix(s: str) -> str:
    """Remove trailing '(...)' color suffixes at the end of a part string."""
    if not s:
        return s
    out = str(s)
    while True:
        m = re.search(r"[（(][^)）]+[)）]$", out)
        if not m:
            break
        out = out[:m.start()]
    return out.strip()


# =====================================================
# COLOR SPLITTER
# =====================================================

COLOR_PAREN_RE = re.compile(r"[（(]([^)）]+)[)）]$")


def split_excel_color(raw_color: str, full_name: str = "") -> List[str]:
    colors: List[str] = []
    raw_color = normalize_color_key(raw_color)

    if not raw_color and full_name:
        m = COLOR_PAREN_RE.search(full_name)
        if m:
            raw_color = normalize_color_key(m.group(1))

    if not raw_color:
        return colors

    if "・" in raw_color:
        main, sub = raw_color.split("・", 1)
        main = normalize_color_key(main)
        sub = normalize_color_key(sub)
        if main:
            colors.append(main)
        if sub:
            sub_norm = normalize_threef_black(sub)
            colors.append("3F黒" if sub_norm == "3F黒" else sub_norm)
    else:
        raw_norm = normalize_threef_black(raw_color)
        colors.append("3F黒" if raw_norm == "3F黒" else raw_color)

    return [c for c in colors if c]


# =====================================================
# YELLOW / HIGHLIGHT DETECTION
# =====================================================

def _color_to_hex(col) -> str:
    if not col:
        return ""
    idx = getattr(col, "indexed", None)
    if idx == 6:
        return "FFFF00"
    rgb = getattr(col, "rgb", None)
    if rgb:
        if isinstance(rgb, str):
            return rgb.upper()
        else:
            try:
                return str(rgb).upper()
            except Exception:
                pass
    val = getattr(col, "value", None)
    if isinstance(val, str):
        return val.upper()
    return ""


def cell_is_yellow(cell) -> bool:
    fill = cell.fill
    if not fill or fill.fill_type in (None, "none"):
        return False
    candidates = [
        getattr(fill, "start_color", None),
        getattr(fill, "end_color", None),
        getattr(fill, "fgColor", None),
        getattr(fill, "bgColor", None),
    ]
    YELLOWS = {"FFFF00", "FFFFFF00", "00FFFF00", "FF00FFFF00"}
    for col in candidates:
        hx = _color_to_hex(col)
        if not hx:
            continue
        short = hx[-6:]
        if short in YELLOWS:
            return True
    return False


def is_row_highlighted(row) -> bool:
    for cell in row[:40]:
        if cell_is_yellow(cell):
            return True
    return False


# =====================================================
# SMALL CONVERSION HELPERS
# =====================================================

def _clean_str(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ["nan", "nat", "none"]:
        return ""
    return s


def _ceil_number(val) -> int:
    if val in (None, ""):
        return 0
    try:
        x = float(val)
        return int(math.ceil(x))
    except Exception:
        return 0


def format_date_mm_dd(v) -> str:
    """Return date as 'MM/DD' from datetime/date/strings (no year)."""
    if isinstance(v, datetime):
        return v.strftime("%m/%d")
    if isinstance(v, date):
        return v.strftime("%m/%d")
    s = str(v or "").strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%m/%d")
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s).strftime("%m/%d")
    except Exception:
        pass
    m = re.search(r"(\d{4})[\-/](\d{1,2})[\-/](\d{1,2})|(\d{1,2})[\-/](\d{1,2})", s)
    if m:
        if m.group(2) and m.group(3):
            return f"{m.group(2).zfill(2)}/{m.group(3).zfill(2)}"
        if m.group(4) and m.group(5):
            return f"{m.group(4).zfill(2)}/{m.group(5).zfill(2)}"
    return s


def format_date_yyyy_mm_dd(v) -> str:
    """Return date as 'YYYY/MM/DD' with no time. Accepts datetime/date/strings."""
    if isinstance(v, datetime):
        return v.strftime("%Y/%m/%d")
    if isinstance(v, date):
        return v.strftime("%Y/%m/%d")
    s = str(v or "").strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y/%m/%d")
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s).strftime("%Y/%m/%d")
    except Exception:
        pass
    m = re.search(r"(\d{4})[\-\/]?(\d{2})[\-\/]?(\d{2})", s)
    if m:
        y, mth, d = m.group(1), m.group(2), m.group(3)
        return f"{y}/{mth}/{d}"
    return s


def _get_row_ct_like_daily(row, headers: Dict[str, int]) -> int:
    idx = headers.get("作業時間(秒)")
    if idx is None:
        return 0
    cell = row[idx].value
    return _ceil_number(cell)


# =====================================================
# NOTION HELPERS
# =====================================================

# =====================================================
# CALENDAR PAGE HELPERS (gap-safe)
# =====================================================

# =====================================================
# CALENDAR HELPERS (gap-safe)
# =====================================================

def _title_to_date(title: str) -> Optional[date]:
    """
    Calendar page title is expected like 'MM/DD/YYYY' (your app uses this).
    Returns a date or None.
    """
    t = (title or "").strip()
    if not t:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(t, fmt).date()
        except Exception:
            pass
    return None


def _get_page_title(page: dict) -> str:
    props = page.get("properties", {}) or {}
    # Common Notion title prop names: "Name", "名前", etc.
    for key, prop in props.items():
        if isinstance(prop, dict) and prop.get("type") == "title":
            frag = prop.get("title") or []
            if frag:
                return (frag[0].get("plain_text") or "").strip()
    return ""


def get_calendar_pages_next_n(n: int = 4, lookahead_days: int = 120) -> List[Tuple[str, str, str]]:
    """
    Return the next available N calendar pages (gap-safe).
    Starts from tomorrow, then searches forward up to lookahead_days
    and returns the next N pages that exist in CALENDAR_DATABASE_ID.

    Output: [(page_id, title, "YYYY-MM-DD"), ...]
    """
    today = date.today()
    start = today + timedelta(days=1)
    end = today + timedelta(days=lookahead_days)

    # Pull candidate pages from Notion. We query without a strict filter on date
    # because your DB may rely on title dates; then we parse titles and filter locally.
    pages = []
    cursor = None
    while True:
        resp = notion_query_database(CALENDAR_DATABASE_ID, start_cursor=cursor)
        pages.extend(resp.get("results", []))
        if not resp.get("has_more"):
            break
        cursor = resp.get("next_cursor")

    candidates: List[Tuple[date, str, str]] = []
    for p in pages:
        title = _get_page_title(p)
        d = _title_to_date(title)
        if not d:
            continue
        if start <= d <= end:
            candidates.append((d, p["id"], title))

    # Sort and take next N
    candidates.sort(key=lambda x: x[0])
    out: List[Tuple[str, str, str]] = []
    for d, pid, title in candidates[: max(0, int(n))]:
        out.append((pid, title, d.isoformat()))
    return out


def get_calendar_pages_next4() -> List[Tuple[str, str, str]]:
    return get_calendar_pages_next_n(4)

def _detect_calendar_date_prop_name() -> str:
    """Detect the Calendar DB's date property name from the schema."""
    try:
        db = _with_backoff(notion.databases.retrieve, database_id=CALENDAR_DATABASE_ID)
        props = db.get("properties", {}) or {}
        for name, info in props.items():
            if isinstance(info, dict) and info.get("type") == "date":
                return name
    except Exception as e:
        log(f"[Calendar] failed to detect date property: {e}")
    # Fallback (common property names)
    for cand in ("日付", "Date", "date"):
        return cand
    return "日付"


def _extract_page_title_from_props(page: dict) -> str:
    """Extract title text from the first 'title' property in a Notion page."""
    props = page.get("properties", {}) or {}
    for _, prop in props.items():
        if isinstance(prop, dict) and prop.get("type") == "title":
            tarr = prop.get("title", []) or []
            return "".join(x.get("plain_text", "") for x in tarr if isinstance(x, dict))
    return ""


def get_calendar_pages_next_n(n: int = 4, lookahead_days: int = 120) -> List[Tuple[str, str, str]]:
    """Return the next available N calendar pages starting from today (gap-safe)."""
    verify_notion_config()
    try:
        n = int(n)
    except Exception:
        n = 4
    if n <= 0:
        return []

    date_prop = _detect_calendar_date_prop_name()

    today = date.today()
    end_day = today + timedelta(days=int(lookahead_days))

    resp = notion_query_database(
        CALENDAR_DATABASE_ID,
        filter={
            "and": [
                {"property": date_prop, "date": {"on_or_after": today.isoformat()}},
                {"property": date_prop, "date": {"on_or_before": end_day.isoformat()}},
            ]
        },
        sorts=[{"property": date_prop, "direction": "ascending"}],
        page_size=100,
    )

    out: List[Tuple[str, str, str]] = []
    for p in (resp.get("results", []) or []):
        pid = p.get("id")
        props = p.get("properties", {}) or {}
        d_str = ((props.get(date_prop, {}) or {}).get("date", {}) or {}).get("start") or ""
        title = _extract_page_title_from_props(p)
        if pid and d_str:
            out.append((pid, title, d_str))
        if len(out) >= n:
            break

    return out


def get_calendar_pages_next4() -> List[Tuple[str, str, str]]:
    """Convenience: next available 4 calendar pages (gap-safe)."""
    return get_calendar_pages_next_n(4)


def get_calendar_pages_today_plus3() -> List[Tuple[str, str, str]]:
    """Backward-compatible alias.

    Historically this returned tomorrow..tomorrow+3.
    Now it returns the next available 4 pages (gap-safe).
    """
    return get_calendar_pages_next4()

def get_all_pages(database_id: str) -> List[dict]:
    results = []
    cursor = None
    while True:
        resp = notion_query_database(database_id, start_cursor=cursor)
        results.extend(resp.get("results", []))
        if not resp.get("has_more"):
            break
        cursor = resp.get("next_cursor")
    return results


def build_parts_map() -> Dict[str, str]:
    verify_notion_config()
    _ensure_cache_dir()
    if _is_cache_fresh(PARTS_CACHE_FILE, CACHE_TTL_SECONDS):
        try:
            with open(PARTS_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f).get("parts_map", {})
        except Exception:
            pass
    pages = get_all_pages(PARTS_DATABASE_ID)
    parts_map = {}
    for p in pages:
        frag = p.get("properties", {}).get("品番", {}).get("title", [])
        if frag:
            name = frag[0].get("plain_text", "").strip()
            if name:
                parts_map[name] = p["id"]
    with open(PARTS_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump({"parts_map": parts_map}, f, ensure_ascii=False, indent=2)
    return parts_map


def get_all_blocks(block_id: str) -> List[dict]:
    blocks = []
    cursor = None
    while True:
        res = _with_backoff(notion.blocks.children.list, block_id=block_id, start_cursor=cursor)
        blocks.extend(res.get("results", []))
        if not res.get("has_more"):
            break
        cursor = res.get("next_cursor")
    return blocks


def _cache_nested_db_id(page_id: str, db_title: str, db_id: str):
    _ensure_cache_dir()
    data = {}
    if os.path.isfile(NESTED_DB_CACHE_FILE):
        try:
            data = json.load(open(NESTED_DB_CACHE_FILE, "r", encoding="utf-8"))
        except Exception:
            data = {}
    data[f"{page_id}:{db_title}"] = {"db_id": db_id, "ts": time.time()}
    json.dump(data, open(NESTED_DB_CACHE_FILE, "w", encoding="utf-8"), ensure_ascii=False, indent=2)


def _load_cached_nested_db_id(page_id: str, db_title: str):
    if not os.path.isfile(NESTED_DB_CACHE_FILE):
        return None
    try:
        data = json.load(open(NESTED_DB_CACHE_FILE, "r", encoding="utf-8"))
        rec = data.get(f"{page_id}:{db_title}")
        if rec and (time.time() - rec["ts"] < CACHE_TTL_SECONDS):
            return rec["db_id"]
    except Exception:
        return None
    return None


def find_nested_databases(page_id: str, target="作業内容") -> List[str]:
    cached = _load_cached_nested_db_id(page_id, target)
    if cached:
        return [cached]
    found: List[str] = []

    def _rec(blocks):
        for b in blocks:
            if b.get("type") == "child_database" and b["child_database"]["title"] == target:
                found.append(b["id"])
                _cache_nested_db_id(page_id, target, b["id"])
            elif b.get("has_children"):
                _rec(get_all_blocks(b["id"]))

    _rec(get_all_blocks(page_id))
    return found


def retrieve_db_ct_is_number(database_id: str) -> bool:
    """Return True if 'c/t 秒' in the database schema is a number; otherwise False."""
    try:
        db = notion.databases.retrieve(database_id=db_id)
        props = db.get("properties", {}) or {}
        p = props.get("c/t 秒")
        if not p:
            # Fallback: try common variants
            for k in ("c/t", "ct", "c/t秒", "c/t  秒", "c/t_sec"):
                if k in props:
                    p = props[k]
                    break
        if not p:
            # Default to number (safer for most of your DBs)
            return True
        return (p.get("type") == "number")
    except Exception:
         # If we cannot read schema (network/auth), assume number to avoid rich_text->number crash
        return True


def build_ct_prop(ct_value, ct_is_number: bool = True):
    """Build Notion property payload for the CT column.

    - If the DB expects Number: {"number": <int>}
    - If the DB expects Rich text: {"rich_text": [{"type":"text","text":{"content":"123"}}]}

    This function is defensive: if ct_is_number is True we always return number.
    """
    try:
        n = int(ct_value or 0)
    except Exception:
        n = 0

    if ct_is_number:
        return {"number": n}

    return {
        "rich_text": [
            {
                "type": "text",
                "text": {"content": str(n)},
            }
        ]
    }

def read_ct_prop_from_page_props(props: dict, is_number: bool) -> int:
    if is_number:
        try:
            return int(props.get("c/t 秒", {}).get("number") or 0)
        except Exception:
            return 0
    rt = props.get("c/t 秒", {}).get("rich_text", []) or []
    try:
        txt = "".join(f.get("text", {}).get("content", "") for f in rt if f.get("type") == "text")
        digits = "".join(ch for ch in txt if ch.isdigit()) or "0"
        return int(digits)
    except Exception:
        return 0


def get_calendar_pages_today_plus3_legacy() -> List[Tuple[str, str, str]]:
    today = date.today()
    dates = [(today + timedelta(days=i)).isoformat() for i in range(5)]
    pages = get_all_pages(CALENDAR_DATABASE_ID)
    out: List[Tuple[str, str, str]] = []
    for r in pages:
        props = r.get("properties", {})
        if "日付" not in props or not props["日付"]["date"]:
            continue
        d = props["日付"]["date"]["start"]
        if d not in dates:
            continue
        title = None
        for k, v in props.items():
            if v.get("type") == "title" and v.get("title"):
                title = v["title"][0]["plain_text"]
                break
        out.append((r["id"], title or d, d))
    out.sort(key=lambda x: x[2])
    return out


# =====================================================
# RICH TEXT HELPERS
# =====================================================

def normalize_qty_str(q: Optional[str]) -> str:
    if q is None:
        return ""
    s = str(q).strip()
    if s == "":
        return ""
    s = s.replace(",", "")
    try:
        f = float(s)
        i = int(f)
        if abs(f - i) < 1e-6:
            return str(i)
        return s
    except ValueError:
        return s


def qty_strings_equal(a: Optional[str], b: Optional[str]) -> bool:
    return normalize_qty_str(a) == normalize_qty_str(b)


def rich_text_is_effectively_empty(rt: List[dict]) -> bool:
    if not rt:
        return True
    for frag in rt:
        if frag.get("type") == "mention":
            return False
        if frag.get("type") == "text":
            content = frag.get("text", {}).get("content", "") or ""
            if content.strip() and content.strip() != "\n":
                return False
    return True


def rich_text_to_lines(rt: List[dict]) -> List[List[dict]]:
    lines: List[List[dict]] = []
    cur: List[dict] = []
    for frag in rt:
        if frag.get("type") == "text" and frag.get("text", {}).get("content") == "\n":
            lines.append(cur)
            cur = []
        else:
            cur.append(frag)
    if cur:
        lines.append(cur)
    return lines


def lines_to_rich_text(lines: List[List[dict]]) -> List[dict]:
    out: List[dict] = []
    for line in lines:
        out.extend(line)
        out.append({"type": "text", "text": {"content": "\n"}})
    if out and out[-1].get("type") == "text" and out[-1]["text"]["content"] == "\n":
        out.pop()
    return out


def parse_parts_lines(parts_rt: List[dict]) -> List[dict]:
    lines = rich_text_to_lines(parts_rt)
    out = []
    for idx, line in enumerate(lines):
        mention_id = None
        part_text_accum = []
        date_token = ""
        has_trial_style = False
        for frag in line:
            t = frag.get("type")
            if t == "mention" and frag.get("mention", {}).get("type") == "page":
                mention_id = frag["mention"]["page"]["id"]
            elif t == "text":
                content = frag.get("text", {}).get("content", "") or ""
                ann = frag.get("annotations", {}) or {}
                if ann.get("bold") and ann.get("color") == "red_background":
                    if re.fullmatch(r"(試作\d+)", content.strip()):
                        has_trial_style = True
                if ann.get("italic") and ann.get("color") in ("blue", "green"):
                    date_token = content.strip()
                if content and content != "\n":
                    part_text_accum.append(content)
        part_text = "".join(part_text_accum).strip()
        out.append(
            {
                "idx": idx,
                "mention_id": mention_id,
                "part_text": part_text,
                "date": date_token,
                "is_app_style": has_trial_style,
            }
        )
    return out


def ensure_parts_has_separator(parts_rt: List[dict]) -> List[dict]:
    if not parts_rt:
        return parts_rt
    last = parts_rt[-1]
    if last.get("type") == "text" and last.get("text", {}).get("content", "") == "\n":
        return parts_rt
    parts_rt.append({"type": "text", "text": {"content": "\n"}})
    return parts_rt


def trim_parts_trailing_newline(parts_rt: List[dict]) -> List[dict]:
    if parts_rt and parts_rt[-1].get("type") == "text" and parts_rt[-1]["text"]["content"] == "\n":
        parts_rt = parts_rt[:-1]
    return parts_rt


def find_part_line_index_strict(parts_rt: List[dict], part_page_id: Optional[str], part_name: str,
                                date_str: str) -> Optional[int]:
    parsed = parse_parts_lines(parts_rt)
    if not parsed:
        return None
    date_str = (date_str or "").strip()
    norm_target = normalize_part_key(part_name or "")
    if part_page_id and date_str:
        for e in parsed:
            if e["mention_id"] == part_page_id and e["date"] == date_str:
                return e["idx"]
    if part_name and date_str:
        for e in parsed:
            if e["date"] == date_str and e["part_text"] == part_name:
                return e["idx"]
    if norm_target and date_str:
        for e in parsed:
            if e["date"] == date_str and normalize_part_key(e["part_text"]) == norm_target:
                return e["idx"]
    if part_page_id:
        for e in parsed:
            if e["mention_id"] == part_page_id:
                return e["idx"]
    if part_name:
        for e in parsed:
            if e["part_text"] == part_name:
                return e["idx"]
    if norm_target:
        for e in parsed:
            if normalize_part_key(e["part_text"]) == norm_target:
                return e["idx"]
    return None


def find_part_line_index(parts_rt: List[dict], part_page_id: Optional[str], part_name: str,
                         date_str: str) -> Optional[int]:
    parsed = parse_parts_lines(parts_rt)
    if not parsed:
        return None
    date_str = (date_str or "").strip()
    norm_target = normalize_part_key(part_name or "")
    if part_page_id and date_str:
        for e in parsed:
            if e["mention_id"] == part_page_id and e["date"] == date_str:
                return e["idx"]
    if part_name and date_str:
        for e in parsed:
            if e["date"] == date_str and e["part_text"] and (part_name in e["part_text"]):
                return e["idx"]
    if norm_target and date_str:
        for e in parsed:
            if e["date"] == date_str and normalize_part_key(e["part_text"]) == norm_target:
                return e["idx"]
    if part_page_id:
        for e in parsed:
            if e["mention_id"] == part_page_id:
                return e["idx"]
    if part_name:
        for e in parsed:
            if e["part_text"] and (part_name in e["part_text"]):
                return e["idx"]
    if norm_target:
        for e in parsed:
            if normalize_part_key(e["part_text"]) == norm_target:
                return e["idx"]
    if len(parsed) == 1:
        return parsed[0]["idx"]
    return None


def append_qty_green_italic(qty_rt: List[dict], qty_value: str) -> List[dict]:
    if not qty_value:
        if APP_SKIP_EMPTY_QTY:
            return qty_rt or []
    norm = normalize_qty_str(qty_value)
    if norm == "" and APP_SKIP_EMPTY_QTY:
        return qty_rt or []
    new_rt = list(qty_rt or [])
    if new_rt and not (new_rt[-1].get("type") == "text" and new_rt[-1]["text"].get("content", "") == "\n"):
        new_rt.append({"type": "text", "text": {"content": "\n"}})
    new_rt.append(
        {
            "type": "text",
            "text": {"content": norm},
            "annotations": {"bold": True, "italic": False, "color": APP_QTY_COLOR},
        }
    )
    return new_rt


def remove_qty_at_index_if_green_italic_with_value(qty_rt: List[dict], index: int, qty_value: str) -> List[dict]:
    from math import inf

    if qty_rt is None:
        return []
    lines = rich_text_to_lines(qty_rt)
    want_norm = normalize_qty_str(qty_value)

    def line_has_app_qty(line_frags: List[dict]) -> bool:
        for f in line_frags:
            if f.get("type") != "text":
                continue
            txt = (f.get("text", {}) or {}).get("content", "").strip()
            txt_norm = normalize_qty_str(txt)
            if not qty_strings_equal(txt_norm, want_norm):
                continue
            ann = (f.get("annotations", {}) or {})
            if APP_QTY_REQUIRE_BOLD and not ann.get("bold"):
                continue
            if ann.get("color") != APP_QTY_COLOR:
                continue
            if ann.get("italic"):
                continue
            return True
        return False

    candidates = [index]
    if APP_QTY_ALLOW_NEIGHBORS:
        candidates.extend([index - 1, index + 1])

    chosen_idx = None
    for cand in candidates:
        if 0 <= cand < len(lines) and line_has_app_qty(lines[cand]):
            chosen_idx = cand
            break

    if chosen_idx is None:
        best_i = None
        best_dist = inf
        for i, ln in enumerate(lines):
            if line_has_app_qty(ln):
                dist = abs(i - (index if index is not None else i))
                if dist < best_dist:
                    best_dist = dist
                    best_i = i
        chosen_idx = best_i

    if chosen_idx is None:
        return qty_rt

    del lines[chosen_idx]

    normalized_lines = []
    prev_blank = False
    for ln in lines:
        is_blank = (len(ln) == 0)
        if is_blank:
            if not prev_blank:
                normalized_lines.append([])
            prev_blank = True
        else:
            normalized_lines.append(ln)
            prev_blank = False

    return lines_to_rich_text(normalized_lines)


def remove_parts_line_at_index(parts_rt: List[dict], index: int) -> List[dict]:
    lines = rich_text_to_lines(parts_rt)
    if 0 <= index < len(lines):
        del lines[index]
    return lines_to_rich_text(lines)


def scan_child_db_cache(nested_db_id: str) -> Dict[str, dict]:
    ct_is_number = retrieve_db_ct_is_number(nested_db_id)
    cache: Dict[str, dict] = {}
    rows = get_all_pages(nested_db_id)
    for r in rows:
        props = r.get("properties", {})
        color_title = props.get("色", {}).get("title", [])
        color = color_title[0]["plain_text"].strip() if color_title else ""
        if not color:
            continue
        parts_entries = parse_parts_lines(props.get("品番", {}).get("rich_text", []) or [])
        ct_val = read_ct_prop_from_page_props(props, ct_is_number)
        cache[color] = {"page": r, "parts": parts_entries, "qty_lines": [], "ct": ct_val}
    return cache


# =====================================================
# DAILY-LIKE CT HELPERS
# =====================================================

CT_SPLIT_MAIN_RATIO = 0.65
CT_SPLIT_3F_RATIO = 0.35


def allocate_ct_for_colors(colors: List[str], total_ct: int) -> Dict[str, int]:
    """Allocate cycle time per color.
    If exactly two colors and the second is '3F黒', allocate 65% to the first (main)
    and the remainder (35%) to '3F黒'. Otherwise, keep legacy behavior (each gets total_ct).
    """
    try:
        total_ct = int(total_ct or 0)
    except Exception:
        total_ct = 0

    norm = [("3F黒" if normalize_threef_black(c) == "3F黒" else c) for c in (colors or [])]
    if len(norm) == 2 and norm[1] == "3F黒":
        main = norm[0]
        main_ct = int(math.ceil(total_ct * CT_SPLIT_MAIN_RATIO))
        sub_ct = max(0, total_ct - main_ct)
        return {main: main_ct, "3F黒": sub_ct}

    return {c: total_ct for c in norm if c}


# =====================================================
# LOT NUMBERING (子品番の正式名称に ロット# を付ける)
# =====================================================

# =====================================================
# LOT NUMBERING (品目名称に ロット# を付ける)
# =====================================================

_LOT_SUFFIX_RE = re.compile(r"(?:\s*ロット\s*#?\s*\d+)\s*$")

def _strip_lot_suffix(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    return _LOT_SUFFIX_RE.sub("", s).strip()


def apply_lot_numbers_to_excel(filepath: str) -> bool:
    """
    If duplicates exist, append ロット# to 品目名称 so rows become distinguishable.

    Duplicates are grouped by:
      (塗装色, 品目名称(base without ロット suffix), 子品番の正式名称)

    Returns True if file was modified.
    """
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        headers = {c.value: i for i, c in enumerate(ws[1])}

        idx_part = headers.get("品目名称")
        idx_full = headers.get("子品番の正式名称")
        idx_color = headers.get("塗装色")

        if idx_part is None:
            return False

        groups = defaultdict(list)  # key -> list[cell_for_part]

        for row in ws.iter_rows(min_row=2):
            part = str(row[idx_part].value or "").strip()
            if not part:
                continue
            base_part = _strip_lot_suffix(part)

            full = ""
            if idx_full is not None:
                full = str(row[idx_full].value or "").strip()

            raw_color = ""
            if idx_color is not None:
                raw_color = str(row[idx_color].value or "").strip()

            # Optional: if 塗装色 empty, try derive from full_name "(...)"
            if (not raw_color) and full:
                m = COLOR_PAREN_RE.search(full)
                if m:
                    raw_color = str(m.group(1) or "").strip()

            key = (raw_color, base_part, full)
            groups[key].append(row[idx_part])  # <-- edit 品目名称 cell

        changed = False

        # Strip existing ロット suffix from 品目名称 (idempotent)
        for key, cells in groups.items():
            for cell in cells:
                cur = str(cell.value or "").strip()
                stripped = _strip_lot_suffix(cur)
                if stripped != cur:
                    cell.value = stripped
                    changed = True

        # Assign lot numbers for duplicates
        for key, cells in groups.items():
            if len(cells) <= 1:
                continue
            for i, cell in enumerate(cells, start=1):
                base = _strip_lot_suffix(str(cell.value or "").strip())
                new_val = f"{base} ロット{i}" if base else f"ロット{i}"
                if str(cell.value or "").strip() != new_val:
                    cell.value = new_val
                    changed = True

        if not changed:
            return False

        fd, tmp = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(tmp)
        shutil.move(tmp, filepath)
        return True

    except Exception as e:
        try:
            log(f"[LOT] apply_lot_numbers_to_excel error: {e}")
        except Exception:
            pass
        return False

# =====================================================
# EXCEL PREPROCESSOR (N93・3F黒 split)
# =====================================================

def preprocess_excel_split_n93_3f(filepath: str):
    """
    Preprocess the Excel file so that any row whose 『塗装色』 is like 'N93・3F黒' or
    'N93・3分艶ﾌﾞﾗｯｸ' is split into two physical rows and the original row is deleted:
      - MAIN row (left of '・') with 65% of 作業時間(秒)
      - '3F黒' row with the remaining 35%
    Also appends (色) to 品目名称 for both rows and ensures 開始日 is a date.
    """
    try:
        wb = load_workbook(filepath)
        ws = wb.active
    except Exception as e:
        log(f"[Preprocess] Failed to open workbook: {e}")
        return

    try:
        wb_vals = load_workbook(filepath, data_only=True)
        ws_vals = wb_vals.active
    except Exception as e:
        wb_vals = None
        ws_vals = None
        log(f"[Preprocess] data_only workbook open failed: {e}")

    headers = {str(c.value).strip() if c.value else "": idx + 1 for idx, c in enumerate(ws[1])}
    col_color = headers.get("塗装色")
    col_ct = headers.get("作業時間(秒)")
    col_date = headers.get("開始日")
    col_part = headers.get("品目名称")

    if not col_color or not col_ct:
        log("[Preprocess] Required headers not found (塗装色 / 作業時間(秒)). Skipped.")
        return

    def _to_float(val) -> float:
        try:
            if val is None or (isinstance(val, str) and val.strip() == ""):
                return 0.0
            return float(val)
        except Exception:
            return 0.0

    def _to_date(val):
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        s = str(val or "").strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
        try:
            return datetime.fromisoformat(s).date()
        except Exception:
            return None

    i = 2
    max_row = ws.max_row
    made_changes = False

    while i <= max_row:
        raw = ws.cell(row=i, column=col_color).value
        color_str = str(raw).strip() if raw is not None else ""

        if "・" in color_str:
            main, sub = color_str.split("・", 1)
            main = (main or "").strip()
            sub = (sub or "").strip()

            if main and normalize_threef_black(sub) == "3F黒":
                raw_ct_val = ws.cell(row=i, column=col_ct).value
                total_ct = None

                if isinstance(raw_ct_val, (int, float)):
                    total_ct = _to_float(raw_ct_val)
                elif isinstance(raw_ct_val, str):
                    try:
                        total_ct = _to_float(raw_ct_val)
                    except Exception:
                        total_ct = None

                if total_ct is None:
                    if ws_vals is not None:
                        total_ct = _to_float(ws_vals.cell(row=i, column=col_ct).value)
                    else:
                        total_ct = _to_float(raw_ct_val)

                main_ct = round(total_ct * 0.65, 1)
                sub_ct = round(total_ct - main_ct, 1)

                row_values = [ws.cell(row=i, column=c).value for c in range(1, ws.max_column + 1)]
                orig_part = row_values[col_part - 1] if col_part else None
                base_part = strip_trailing_color_suffix(orig_part) if orig_part else None

                log(f"[Preprocess] Split '{main}・{sub}' part='{orig_part}' total_ct={total_ct} -> main={main_ct}, sub={sub_ct} (row {i})")

                ws.insert_rows(i + 1, amount=2)

                for c in range(1, ws.max_column + 1):
                    ws.cell(row=i + 1, column=c, value=row_values[c - 1])
                ws.cell(row=i + 1, column=col_color, value=main)
                ws.cell(row=i + 1, column=col_ct, value=main_ct)
                if col_part and base_part is not None:
                    ws.cell(row=i + 1, column=col_part, value=f"{base_part}({main})")

                for c in range(1, ws.max_column + 1):
                    ws.cell(row=i + 2, column=c, value=row_values[c - 1])
                ws.cell(row=i + 2, column=col_color, value="3F黒")
                ws.cell(row=i + 2, column=col_ct, value=sub_ct)
                if col_part and base_part is not None:
                    ws.cell(row=i + 2, column=col_part, value=f"{base_part}(3F黒)")

                if col_date:
                    dobj = _to_date(row_values[col_date - 1])
                    if dobj:
                        dcell_main = ws.cell(row=i + 1, column=col_date, value=dobj)
                        dcell_sub = ws.cell(row=i + 2, column=col_date, value=dobj)
                        dcell_main.number_format = "yyyy/mm/dd"
                        dcell_sub.number_format = "yyyy/mm/dd"

                ws.delete_rows(i, amount=1)

                made_changes = True
                max_row += 1
                i += 2
                continue

        i += 1

    if made_changes:
        temp_path = filepath.replace(".xlsx", "_temp.xlsx")
        try:
            wb.save(temp_path)
            os.replace(temp_path, filepath)
            log("[Preprocess] Split N93・3F黒 rows and saved.")
        except Exception:
            try:
                wb.save(filepath)
                log("[Preprocess] Saved changes directly to workbook.")
            except Exception as e:
                log(f"[Preprocess] Failed to save workbook: {e}")
    else:
        log("[Preprocess] No target rows (N93・3F黒 / 3分艶ﾌﾞﾗｯｸ) were found.")


# =====================================================
# DAILY GENERATOR HELPERS (Excel → grouped data)
# =====================================================

TRIAL_CODE_RE = re.compile(r"(試作\d+)")


def find_excel_for_specific_date(path: str, target_date: date) -> Optional[str]:
    target_md = target_date.strftime("%m%d")
    if not os.path.isdir(path):
        return None
    candidates = [f for f in os.listdir(path) if f.startswith(f"{target_md}_塗装") and f.endswith(".xlsx")]
    if not candidates:
        return None
    latest = max([os.path.join(path, f) for f in candidates], key=os.path.getmtime)
    return latest


def read_excel_data_pandas(filepath: str) -> pd.DataFrame:
    cols = ["開始日", "塗装色", "子品番の正式名称", "品目名称", "試作番号", "完成品数", "作業時間(秒)"]
    return pd.read_excel(filepath, usecols=cols, engine="openpyxl")


def group_data_daily(df: pd.DataFrame) -> Dict[str, Dict[str, List[str]]]:
    """Group Excel rows by color for Daily Workflow Generator.

    NEW rule (Reo):
      - Duplicates are based on (色, 試作番号, 品目名称).
      - If all 子品番の正式名称 values match (or are all empty): DO NOT MERGE.
        Keep each Excel row as its own line in Notion (数量 is kept as-is per row).
        Total c/t is summed by adding each row’s c/t.
      - If 子品番の正式名称 differs within the duplicates: MERGE into one line.
        Keep 数量 as the first non-empty value, and sum c/t.

    Notes:
      - Mixed colors like 'N93・3F黒' still allocate ct via allocate_ct_for_colors.
    """

    groups: Dict[str, Dict[str, List[str] | str | int]] = defaultdict(
        lambda: {"品番": [], "数量": [], "作業時間(秒)": 0}
    )

    def _clean_fn(x: str) -> str:
        return (x or "").strip()

    def _qty_first_non_empty(items: List[dict]) -> str:
        for it in items:
            q = it.get("qty")
            if str(q or "").strip():
                return str(q).strip()
        return ""

    # per-color accumulator: color -> key(part_key, trial) -> list of row-items
    per_color: Dict[str, Dict[Tuple[str, str], List[dict]]] = defaultdict(lambda: defaultdict(list))

    for _, row in df.iterrows():
        raw_color = _clean_str(row.get("塗装色", ""))

        full_name = _clean_str(row.get("子品番の正式名称", ""))
        if not raw_color:
            m = COLOR_PAREN_RE.search(full_name)
            if m:
                raw_color = _clean_str(m.group(1))
        if not raw_color:
            continue

        # Resolve colors_to_use using the same logic as before
        colors_to_use: List[str] = []
        if "・" in raw_color:
            main, sub = raw_color.split("・", 1)
            colors_to_use.append(_clean_str(main))
            sub = _clean_str(sub)
            if "3分艶ﾌﾞﾗｯｸ" in sub:
                colors_to_use.append("3F黒")
            elif sub:
                colors_to_use.append(sub)
        else:
            colors_to_use.append(raw_color)

        colors_to_use = ["3F黒" if c == "3分艶ﾌﾞﾗｯｸ" else c for c in colors_to_use]

        part_name = _clean_str(row.get("品目名称", ""))
        if not part_name:
            continue

        trial = _clean_str(row.get("試作番号", ""))
        display = f"{trial}・{part_name}" if trial else part_name

        qty_raw = row.get("完成品数", 0)
        qty_str = str(_ceil_number(qty_raw))

        ct_row = _ceil_number(row.get("作業時間(秒)", 0))

        # Allocate CT for mixed colors
        if "・" in raw_color:
            ct_alloc = allocate_ct_for_colors(colors_to_use, ct_row)
        else:
            ct_alloc = {c: ct_row for c in colors_to_use}

        for color in colors_to_use:
            color = _clean_str(color)
            if not color:
                continue

            color_key = normalize_color_key(color)
            part_key = normalize_part_key(part_name)
            key = (part_key, trial)  # keep trial separation

            per_color[color_key][key].append(
                {
                    "display": display,
                    "full_name": full_name,
                    "qty": qty_str,
                    "ct": int(ct_alloc.get(color, ct_row) or 0),
                }
            )

    # Finalize output structure
    for color, items_by_key in per_color.items():
        for _key, items in items_by_key.items():
            # distinct non-empty full names
            fset = {_clean_fn(it.get("full_name", "")) for it in items if _clean_fn(it.get("full_name", ""))}

            if len(fset) <= 1:
                # DO NOT MERGE: keep one line per original Excel row
                for it in items:
                    groups[color]["品番"].append(it.get("display") or "")
                    groups[color]["数量"].append(str(it.get("qty") or ""))
                    groups[color]["作業時間(秒)"] += int(it.get("ct") or 0)
            else:
                # MERGE: different 子品番の正式名称 exist
                qty0 = _qty_first_non_empty(items)
                ct_sum = sum(int(it.get("ct") or 0) for it in items)
                groups[color]["品番"].append(items[0].get("display") or "")
                groups[color]["数量"].append(qty0)
                groups[color]["作業時間(秒)"] += int(ct_sum)

        groups[color]["作業時間(秒)"] = str(int(groups[color]["作業時間(秒)"]))

    return groups



# Cache for user choices when multiple part candidates exist for the same text
_PART_CHOICE_CACHE: Dict[str, str] = {}


def _ask_user_to_choose_part_key(part_text: str, candidates: List[str]) -> Optional[str]:
    """
    When multiple parts_map keys could match a given part_text, ask the user which one to use.
    Returns the chosen key, or None if the dialog is cancelled or Tk is unavailable.
    """
    # If we already chose for this exact part_text in this run, reuse it.
    if part_text in _PART_CHOICE_CACHE and _PART_CHOICE_CACHE[part_text] in candidates:
        return _PART_CHOICE_CACHE[part_text]

    # If Tk/messagebox is not available (e.g. non-GUI context), fall back to first candidate
    if tk is None or messagebox is None:
        return candidates[0] if candidates else None

    # Build a simple modal dialog with radio buttons
    root = tk._default_root
    if root is None:
        # No default root -> cannot safely show UI, fall back
        return candidates[0] if candidates else None

    dialog = tk.Toplevel(root)
    dialog.title("部品候補の選択")
    dialog.transient(root)
    dialog.grab_set()

    msg = f"「{part_text}」に対して複数の部品候補があります。\n使用する品番を選んでください。"
    tk.Label(dialog, text=msg, justify="left").pack(padx=12, pady=(12, 4), anchor="w")

    var = tk.StringVar(dialog)

    # List radio buttons for each candidate
    for k in candidates:
        rb = tk.Radiobutton(dialog, text=k, value=k, variable=var, anchor="w", justify="left")
        rb.pack(fill="x", padx=20, pady=1, anchor="w")

    chosen = {"value": None}

    def on_ok():
        val = var.get()
        if not val:
            messagebox.showwarning("選択", "品番を選んでください。", parent=dialog)
            return
        chosen["value"] = val
        dialog.destroy()

    def on_cancel():
        dialog.destroy()

    btn_frame = tk.Frame(dialog)
    btn_frame.pack(pady=(8, 12))
    tk.Button(btn_frame, text="OK", width=10, command=on_ok).pack(side="left", padx=4)
    tk.Button(btn_frame, text="キャンセル", width=10, command=on_cancel).pack(side="left", padx=4)

    dialog.wait_window()

    if not chosen["value"]:
        # User cancelled -> fall back to first if available
        return candidates[0] if candidates else None

    _PART_CHOICE_CACHE[part_text] = chosen["value"]
    return chosen["value"]


def _append_part_with_mention_and_modifier_auto(rich_list: List[dict], part_text: str, parts_map: Dict[str, str]):
    """
    Append a part mention (linked to the Parts DB) plus any modifier text.

    Matching priority:
    1) Exact title match in parts_map.
    2) Longest title where part_text.startswith(title).
    3) If multiple longest candidates exist, show a dialog and let the user choose.
    """
    if not part_text:
        return

    # 1) Exact match first (highest precision)
    if part_text in parts_map:
        best_key = part_text
    else:
        # 2) Collect all keys where part_text starts with the key
        max_len = -1
        candidates: List[str] = []
        for k in parts_map.keys():
            if part_text.startswith(k):
                L = len(k)
                if L > max_len:
                    max_len = L
                    candidates = [k]
                elif L == max_len:
                    candidates.append(k)

        if not candidates:
            best_key = None
        elif len(candidates) == 1:
            best_key = candidates[0]
        else:
            # 3) Multiple longest candidates → ask the user which one to use
            log(f"[Link] Multiple part candidates for '{part_text}': {candidates}")
            chosen = _ask_user_to_choose_part_key(part_text, candidates)
            best_key = chosen

    if best_key and best_key in parts_map:
        # Insert mention for the chosen part page
        rich_list.append(
            {
                "type": "mention",
                "mention": {"type": "page", "page": {"id": parts_map[best_key]}},
            }
        )
        # Any remaining suffix text (e.g. ' (黒)', modifiers, etc.) is kept as plain text
        modifier = part_text[len(best_key):]
        if modifier:
            rich_list.append({"type": "text", "text": {"content": modifier}})
    else:
        # Fallback: no reliable match → just plain text
        rich_list.append({"type": "text", "text": {"content": part_text}})


def create_row_in_nested_db_auto(
    nested_db_id,
    color,
    display_names,
    finish_qty,
    total_cycle_time,
    parts_map,
):
    # Determine the actual property type of 'c/t 秒' on this nested DB (Number vs Rich text)
    try:
        ct_is_number = retrieve_db_ct_is_number(nested_db_id)
    except Exception:
        ct_is_number = True

    # Normalize total_cycle_time into an int (handle None/NaN/str)
    try:
        if total_cycle_time is None:
            ct_val_int = 0
        else:
            s = str(total_cycle_time).strip()
            if s.lower() in ("", "nan", "none"):
                ct_val_int = 0
            else:
                ct_val_int = int(math.ceil(float(s)))
    except Exception:
        ct_val_int = 0
    rich_text_parts = []
    for p in display_names:
        if "・" in p:
            maybe_trial, part = p.split("・", 1)
        else:
            maybe_trial, part = "", p
        trial_match = TRIAL_CODE_RE.fullmatch(maybe_trial)
        if trial_match:
            rich_text_parts.append(
                {
                    "type": "text",
                    "text": {"content": trial_match.group(1)},
                    "annotations": {"bold": True, "color": "red_background"},
                }
            )
            rich_text_parts.append({"type": "text", "text": {"content": "・"}})
        elif maybe_trial:
            rich_text_parts.append({"type": "text", "text": {"content": maybe_trial}})
            rich_text_parts.append({"type": "text", "text": {"content": "・"}})

        _append_part_with_mention_and_modifier_auto(rich_text_parts, part, parts_map)
        rich_text_parts.append({"type": "text", "text": {"content": "\n"}})

    if rich_text_parts and rich_text_parts[-1].get("text", {}).get("content") == "\n":
        rich_text_parts.pop()

    qty_rich = []
    for i, q in enumerate(finish_qty):
        qty_rich.append(
            {
                "type": "text",
                "text": {"content": q + ("\n" if i < len(finish_qty) - 1 else "")},
                "annotations": {"bold": True, "color": "orange"},
            }
        )

    _t0 = int(time.time() * 1000)
    log(f"[Notion] START pages.create color={color} db={nested_db_id}")

    # --- normalize ct value (int) ---
    try:
        if total_cycle_time is None:
            ct_val_int = 0
        else:
            s = str(total_cycle_time).strip()
            if s.lower() in ("", "nan", "none"):
                ct_val_int = 0
            else:
                ct_val_int = int(math.ceil(float(s)))
    except Exception:
        ct_val_int = 0

    # Build CT payload using detected schema first
    ct_prop = build_ct_prop(ct_val_int, ct_is_number)

    def _create_with_ct(ct_payload):
        return _with_backoff(
            notion.pages.create,
            parent={"database_id": nested_db_id},
            icon={"type": "emoji", "emoji": "⚙️"},
            properties={
                "色": {"title": [{"text": {"content": color}}]},
                "品番": {"rich_text": rich_text_parts},
                "数量": {"rich_text": qty_rich},
                "c/t 秒": ct_payload,
                "詳細": {"select": {"name": "ライン"}},
            },
        )

    try:
        new_page = _create_with_ct(ct_prop)
        return new_page

    except APIResponseError as e:
        msg = str(e)
        low = msg.lower()

        # If Notion expects a Number, force Number payload and retry once.
        if ("c/t" in low or "c/t 秒" in msg) and ("expected to be number" in low or "is expected to be number" in low):
            log("[Notion] CT type mismatch; retrying with Number payload for 'c/t 秒'.")
            new_page = _create_with_ct({"number": ct_val_int})
            return new_page

        # If Notion expects rich_text, force rich_text payload and retry once.
        if ("c/t" in low or "c/t 秒" in msg) and ("expected to be rich_text" in low or "is expected to be rich_text" in low):
            log("[Notion] CT type mismatch; retrying with rich_text payload for 'c/t 秒'.")
            new_page = _create_with_ct({"rich_text": [{"type": "text", "text": {"content": str(ct_val_int)}}]})
            return new_page

        raise

def get_all_pages(database_id: str, **kwargs) -> List[dict]:
    """Fetch all pages from a Notion database (with pagination)."""
    results = []
    start_cursor = None
    while True:
        resp = notion_query_database(database_id, start_cursor=start_cursor, **kwargs)
        results.extend(resp.get("results", []))
        if not resp.get("has_more"):
            break
        start_cursor = resp.get("next_cursor")
    return results