"""
workflow_manager.py
-------------------

Core logic + shared state for the Workflow Manager feature.

- Keeps track of:
    selected_page_id / selected_page_title / selected_file
    wf_progress flags
    checkbox_vars (MB/OB checkboxes per part)
    color_checkbox_icon_map (for group toggle)

- Exposes:
    highlight_and_sync()   ← main sync function (Excel ⇄ Notion)
    toggle_sync_for_color_group(color_key)

UI screens (main menu, calendar selector, Excel selector, product list UI) will
call these and manipulate the shared state.

This module depends on `logic.py` for all heavy Notion/Excel helpers and on
`layout_themes.py` for theme colors.
"""

from __future__ import annotations

import os
import tempfile
import shutil
from collections import defaultdict
from typing import Dict, List, Optional, Tuple
from datetime import datetime

import tkinter as tk
from tkinter import ttk, messagebox

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from layout_themes import THEMES, CURRENT_THEME
import logic
from logic import (
    preprocess_excel_split_n93_3f,
    is_row_highlighted,
    COLOR_PAREN_RE,
    normalize_color_key,
    normalize_part_key,
    split_excel_color,
    allocate_ct_for_colors,
    _get_row_ct_like_daily,
    format_date_mm_dd,
    find_nested_databases,
    scan_child_db_cache,
    build_parts_map,
    retrieve_db_ct_is_number,
    read_ct_prop_from_page_props,
    parse_parts_lines,
    ensure_parts_has_separator,
    trim_parts_trailing_newline,
    append_qty_green_italic,
    find_part_line_index,
    remove_parts_line_at_index,
    remove_qty_at_index_if_green_italic_with_value,
    rich_text_is_effectively_empty,
    build_ct_prop,
    notion,
)

# Calendar helper (gap-safe): next available 4 days/pages
def get_calendar_pages_next4():
    return logic.get_calendar_pages_next4()

# =====================================================
# GLOBAL UI STATE SHARED WITH APP
# =====================================================

# These will be set / read by your app_main UI code.
selected_page_id: Optional[str] = None
selected_page_title: Optional[str] = None
selected_file: Optional[str] = None

# These widget references are optional; app_main can assign them after creation.
main_frame: Optional[ttk.Frame] = None
status_label: Optional[ttk.Label] = None
log_console: Optional[tk.Text] = None
progress: Optional[ttk.Progressbar] = None
wf_instruction_label: Optional[ttk.Label] = None

# (MB_var, AC_var, OB_var, trial, part, color, qty, ct, date)
checkbox_vars: List[Tuple[tk.BooleanVar, tk.BooleanVar, tk.BooleanVar, str, str, str, int, str]] = []

# Map normalized color -> list of (MB_var, header_icon_label, row_icon_label) for group toggling
color_checkbox_icon_map: Dict[str, List[Tuple[tk.BooleanVar, ttk.Label, ttk.Label]]] = {}

# Workflow step progress flags
wf_progress = {
    "calendar_selected": False,
    "excel_selected": False,
    "products_loaded": False,
}


# =====================================================
# UI-LEVEL HELPERS
# =====================================================

def clear_frame(frame: tk.Frame):
    """Destroy all children in a frame."""
    for w in frame.winfo_children():
        w.destroy()


def log(text: str):
    """Append a timestamped log line into the UI console (if present)."""
    ts = datetime.now().strftime("[%H:%M:%S]")
    msg = f"{ts} {text}"
    # UI text widget
    if log_console and log_console.winfo_exists():
        log_console.insert(tk.END, msg + "\n")
        log_console.see(tk.END)
    # Also forward to backend logger in logic.py (if handler set)
    try:
        logic.log(msg)
    except Exception:
        # fallback: ignore if logic.log not wired
        pass


def set_status(text: str, color: str | None = None):
    """Update the bottom status label (if present)."""
    global status_label
    if not status_label or not status_label.winfo_exists():
        return
    if color is None:
        color = THEMES[CURRENT_THEME]["text"]
    status_label.config(text=text, foreground=color)
    status_label.update_idletasks()


def set_progress(value: Optional[int] = None, mode: str = "determinate"):
    """Control the global progress bar (if present)."""
    global progress
    if not progress or not progress.winfo_exists():
        return
    if mode == "indeterminate":
        progress.config(mode="indeterminate")
        progress.start(12)
    else:
        progress.stop()
        progress.config(mode="determinate")
        if value is not None:
            progress["value"] = value


def toggle_sync_for_color_group(color_key: str):
    """
    Toggle all main 'sync' checkboxes (MB) for a given color group.
    If any are unchecked -> check all. If all are checked -> uncheck all.
    Also updates the emoji labels accordingly (header and row icons).
    """
    target = normalize_color_key(color_key)
    pairs = color_checkbox_icon_map.get(target, [])
    if not pairs:
        return

    any_unchecked = any(not v.get() for (v, _, _) in pairs)
    new_val = True if any_unchecked else False

    for v, header_lbl, row_icon in pairs:
        v.set(new_val)
        try:
            header_lbl.config(text="✅" if new_val else "⬜")
        except Exception:
            pass
        try:
            row_icon.config(text="✅" if new_val else "⬜")
        except Exception:
            pass



# =====================================================
# EXCEL PRODUCT PARSER (for checkbox_vars)
# =====================================================

def load_products_from_excel() -> List[Tuple[tk.BooleanVar, tk.BooleanVar, tk.BooleanVar, str, str, str, int, str]]:
    """
    Parse the currently selected Excel file and populate checkbox_vars.

    Returns
    -------
    List of tuples:
        (MB_var, OB_var, trial, part, color, qty, ct, date)
    """
    global selected_file, checkbox_vars, color_checkbox_icon_map

    if not selected_file:
        messagebox.showwarning("Warning", "先に Excel ファイルを選択してください。")
        return []
    
    # Add ロット# to duplicates so UI/Notion lines become distinguishable
    try:
        logic.apply_lot_numbers_to_excel(selected_file)
    except Exception as e:
        log(f"[LOT] WF lot numbering failed: {e}")

    # Clear previous state
    checkbox_vars.clear()
    color_checkbox_icon_map.clear()

    wb = load_workbook(selected_file)
    ws = wb.active
    headers = {c.value: i for i, c in enumerate(ws[1])}

    # Data-only workbook for CT
    try:
        wb_vals = load_workbook(selected_file, data_only=True)
        ws_vals = wb_vals.active
    except Exception:
        wb_vals = None
        ws_vals = None

    entries: List[Tuple[str, str, str, str, str, int, str, bool]] = []  # (trial, part, full_name, color, qty, ct, date, row_is_yellow)

    for row in ws.iter_rows(min_row=2):
        raw_color = ""
        if headers.get("塗装色") is not None and row[headers["塗装色"]].value:
            raw_color = str(row[headers["塗装色"]].value).strip()

        part = ""
        if headers.get("品目名称") is not None and row[headers["品目名称"]].value:
            part = str(row[headers["品目名称"]].value).strip()
        if not part:
            continue

        full_name = ""
        if "子品番の正式名称" in headers and row[headers["子品番の正式名称"]].value:
            full_name = str(row[headers["子品番の正式名称"]].value).strip()

        trial = ""
        if headers.get("試作番号") is not None and row[headers["試作番号"]].value:
            trial = str(row[headers["試作番号"]].value).strip()

        qty = ""
        if headers.get("完成品数") is not None and row[headers["完成品数"]].value:
            qty = str(row[headers["完成品数"]].value).strip()

        d_val = row[headers["開始日"]].value if headers.get("開始日") is not None else None
        d_str = format_date_mm_dd(d_val)

        # CT: prefer data_only workbook if available
        if ws_vals is not None and headers.get("作業時間(秒)") is not None:
            try:
                row_idx_1 = row[0].row
                col_idx_1 = headers["作業時間(秒)"] + 1
                ct_val = logic._ceil_number(ws_vals.cell(row=row_idx_1, column=col_idx_1).value)
            except Exception:
                ct_val = _get_row_ct_like_daily(row, headers)
        else:
            ct_val = _get_row_ct_like_daily(row, headers)

        colors = split_excel_color(raw_color, full_name)
        # --- Excel yellow highlight detection (inserted here) ---
        row_is_yellow = is_row_highlighted(row)
        # --------------------------------------------------------
        if not colors:
            colors = [normalize_color_key(raw_color)]

        for color in colors:
            entries.append(
                (
                    trial,
                    part,
                    (full_name or "").strip(),
                    normalize_color_key(color),
                    qty,
                    ct_val,
                    d_str,
                    row_is_yellow,
                )
            )

        # Group by (color, 品目名称, 試作番号).
    # NEW RULE (Reo):
    # - If all 子品番の正式名称 values match (or all empty): DO NOT MERGE → keep separate rows.
    # - If 子品番の正式名称 differs: MERGE → keep qty first non-empty, sum ct.

    def _clean_fn(x: str) -> str:
        return (x or "").strip()

    def _qty_first_non_empty(items: list[tuple]) -> str:
        for _trial, _part, _fn, _color, _qty, _ct, _date, _yellow in items:
            if str(_qty or "").strip():
                return str(_qty).strip()
        return ""

    grouped_items: Dict[Tuple[str, str, str], List[Tuple[str, str, str, str, str, int, str, bool]]] = {}
    for t, p, fn, c, q, ct, d, y in entries:
        key = (normalize_color_key(c), normalize_part_key(p), (t or "").strip())
        grouped_items.setdefault(key, []).append((t, p, fn, c, q, ct, d, y))

    out: List[Tuple[tk.BooleanVar, tk.BooleanVar, tk.BooleanVar, str, str, str, int, str]] = []

    for key, items in grouped_items.items():
        fset = {_clean_fn(i[2]) for i in items if _clean_fn(i[2])}

        if len(fset) <= 1:
            # DO NOT MERGE: one UI entry per Excel row
            for (trial, part, full_name, color, qty, ct_val, d_str, was_yellow) in items:
                mb = tk.BooleanVar(value=bool(was_yellow))
                ac = tk.BooleanVar(value=False)
                ob = tk.BooleanVar(value=False)
                out.append((mb, ac, ob, trial, part, normalize_color_key(color), str(qty).strip(), int(ct_val or 0), d_str))
        else:
            # MERGE: different 子品番の正式名称 exist
            trial0 = items[0][0]
            part0 = items[0][1]
            color0 = normalize_color_key(items[0][3])
            date0 = ""
            for i in items:
                if i[6]:
                    date0 = i[6]
                    break
            qty0 = _qty_first_non_empty(items)
            ct_sum = sum(int(i[5] or 0) for i in items)
            was_yellow_any = any(bool(i[7]) for i in items)

            mb = tk.BooleanVar(value=bool(was_yellow_any))
            ac = tk.BooleanVar(value=False)
            ob = tk.BooleanVar(value=False)
            out.append((mb, ac, ob, trial0, part0, color0, qty0, int(ct_sum), date0))

    checkbox_vars.clear()
    checkbox_vars.extend(out)
    return out

def _debug_list_child_databases(page_id: str) -> list[tuple[str, str]]:
    """Best-effort: list child databases under a Notion page for troubleshooting."""
    out: list[tuple[str, str]] = []
    try:
        cursor = None
        while True:
            kwargs = {"block_id": page_id}
            if cursor:
                kwargs["start_cursor"] = cursor
            resp = logic._with_backoff(notion.blocks.children.list, **kwargs)
            results = resp.get("results") or []
            for b in results:
                try:
                    if b.get("type") == "child_database":
                        out.append((b.get("id", ""), b.get("child_database", {}).get("title", "")))
                except Exception:
                    continue
            cursor = resp.get("next_cursor")
            if not cursor:
                break
    except Exception:
        return out
    return out

def highlight_and_sync():
    """
    Main Workflow Manager sync routine.

    Key fixes:
    - Validate Notion child DB BEFORE saving Excel.
    - Only save Excel AFTER Notion updates succeed.
    - Wrap Notion writes with logic._with_backoff so failures are visible.
    """
    global selected_file, selected_page_id

    if not selected_file:
        messagebox.showwarning("Warning", "先にExcelファイルを選択してください。")
        return

    if not selected_page_id:
        messagebox.showwarning("Warning", "カレンダーページを先に選択してください。")
        return

    # ✅ Validate Notion child DB FIRST (before Excel save)
    nested = find_nested_databases(selected_page_id, "作業内容")
    if not nested:
        set_status("作業内容データベースが見つかりません", THEMES[CURRENT_THEME]["status_warn"])
        dbs = _debug_list_child_databases(selected_page_id)
        db_lines = "\n".join([f"- {t or '(no title)'}" for (_id, t) in dbs]) if dbs else "(child_databaseなし)"
        log("[SYNC] No nested DB '作業内容' found. Notion sync aborted. Excel will NOT be saved.")
        log(f"[SYNC] Detected child databases:\n{db_lines}")
        messagebox.showerror(
            "同期できません",
            "選択したカレンダーページの中に『作業内容』(子データベース) が見つかりません。\n\n"
            "そのため、Notionへ同期できません（Excelの変更も保存しません）。\n\n"
            f"検出された子データベース:\n{db_lines}\n\n"
            "対策: Notionのその日付ページの中に『作業内容』という名前の子データベースを作成するか、\n"
            "データベース名を『作業内容』に合わせてください。"
        )
        return

    nested_id = nested[0]
    ct_is_number = retrieve_db_ct_is_number(nested_id)

    # 1) Preprocess N93・3F黒 rows
    try:
        preprocess_excel_split_n93_3f(selected_file)
    except Exception as e:
        log(f"[Manual Sync] Preprocess error: {e}")

    wb = load_workbook(selected_file)
    ws = wb.active
    headers = {c.value: i for i, c in enumerate(ws[1])}

    # Open data-only workbook so formulas yield computed values
    try:
        wb_vals = load_workbook(selected_file, data_only=True)
        ws_vals = wb_vals.active
    except Exception:
        wb_vals = None
        ws_vals = None

    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    empty_fill = PatternFill()

    # 2) Build Excel index keyed by (color_key, part_key)
    excel_index: Dict[Tuple[str, str], dict] = {}

    for row in ws.iter_rows(min_row=2):
        raw_color = ""
        if headers.get("塗装色") is not None and row[headers["塗装色"]].value:
            raw_color = str(row[headers["塗装色"]].value).strip()

        part = ""
        if headers.get("品目名称") is not None and row[headers["品目名称"]].value:
            part = str(row[headers["品目名称"]].value).strip()

        if not part:
            continue

        full_name = ""
        if "子品番の正式名称" in headers and row[headers["子品番の正式名称"]].value:
            full_name = str(row[headers["子品番の正式名称"]].value).strip()

        derived_color = ""
        if not raw_color and full_name:
            m = COLOR_PAREN_RE.search(full_name)
            if m:
                derived_color = normalize_color_key(m.group(1))

        trial = ""
        if headers.get("試作番号") is not None and row[headers["試作番号"]].value:
            trial = str(row[headers["試作番号"]].value).strip()

        qty = ""
        if headers.get("完成品数") is not None and row[headers["完成品数"]].value:
            qty = str(row[headers["完成品数"]].value).strip()

        d_val = row[headers["開始日"]].value if headers.get("開始日") is not None else None
        d_str = format_date_mm_dd(d_val)

        if ws_vals is not None and headers.get("作業時間(秒)") is not None:
            try:
                row_idx_1 = row[0].row
                col_idx_1 = headers["作業時間(秒)"] + 1
                ct_val = logic._ceil_number(ws_vals.cell(row=row_idx_1, column=col_idx_1).value)
            except Exception:
                ct_val = _get_row_ct_like_daily(row, headers)
        else:
            ct_val = _get_row_ct_like_daily(row, headers)

        row_is_yellow = is_row_highlighted(row)

        colors_to_process = split_excel_color(raw_color, full_name)
        if not colors_to_process:
            colors_to_process = [normalize_color_key(raw_color)]

        part_key = normalize_part_key(part)

        for c in colors_to_process:
            color_key = normalize_color_key(c)
            key = (color_key, part_key)

            rec = excel_index.setdefault(
                key,
                {
                    "rows": [],
                    "was_highlighted_before": False,
                    "ct_total": 0,
                },
            )
            rec["rows"].append(row)

            is_mixed = ("・" in raw_color) or ("・" in derived_color)
            if is_mixed:
                ct_alloc = allocate_ct_for_colors(colors_to_process, ct_val)
                rec["ct_total"] += ct_alloc.get(color_key, ct_val)
            else:
                rec["ct_total"] += ct_val

            if row_is_yellow:
                rec["was_highlighted_before"] = True

    # 3) Validate override combos and build action list
    invalid_overrides = []
    override_list = []
    for var, ac_var, override_var, trial, part, color, qty, ct_ui, d in checkbox_vars:
        if override_var.get() and not var.get():
            invalid_overrides.append(f"{color} - {part}")
        if override_var.get() and var.get():
            override_list.append(f"{color} - {part}")

    if invalid_overrides:
        messagebox.showwarning(
            "注意",
            "『上書き』のみが選択されています。\n"
            "次の項目でメインのチェックが外れています:\n\n"
            + "\n".join("・" + x for x in invalid_overrides)
            + "\n\nメインのチェックボックスにもチェックを入れてください。"
        )
        return

    if override_list:
        msg = (
            "次の製品で『上書き』が選択されています。\n"
            "（Excelの状態を無視して同期します）\n\n"
            + "\n".join("・" + x for x in override_list)
            + "\n\nこのまま続行しますか？"
        )
        if not messagebox.askokcancel("上書き確認", msg):
            return

    db_cache = scan_child_db_cache(nested_id)
    existing_by_color = {c: info["page"] for c, info in db_cache.items()}

    actions: List[dict] = []

    for var, ac_var, override_var, trial, part, color, qty, ct_ui, d in checkbox_vars:
        effective_part = part
        if ac_var.get():
            effective_part = f"{part}({color})"

        color_key = normalize_color_key(color)
        rec = excel_index.get((color_key, normalize_part_key(part)))

        was_pre = rec["was_highlighted_before"] if rec else False
        excel_ct = rec["ct_total"] if rec else 0

        is_override = bool(override_var.get())
        chosen_ct = logic._ceil_number(ct_ui)

        if var.get():
            if is_override:
                actions.append({"op": "add", "color": color_key, "trial": trial, "part": effective_part, "qty": qty, "ct": chosen_ct, "date": d, "override": True})
            else:
                if was_pre:
                    continue
                if rec and not rec["was_highlighted_before"]:
                    for r in rec["rows"]:
                        for cell in r[:10]:
                            cell.fill = yellow
                    rec["was_highlighted_before"] = True
                actions.append({"op": "add", "color": color_key, "trial": trial, "part": effective_part, "qty": qty, "ct": chosen_ct, "date": d})
        else:
            if rec and rec["was_highlighted_before"]:
                for r in rec["rows"]:
                    for cell in r[:10]:
                        cell.fill = empty_fill
                rec["was_highlighted_before"] = False
                actions.append({"op": "remove", "color": color_key, "trial": trial, "part": effective_part, "qty": qty, "ct": chosen_ct, "date": d})

    actions_by_color: Dict[str, List[dict]] = defaultdict(list)
    for act in actions:
        actions_by_color[act["color"].strip()].append(act)

    # ✅ Notion sync first (with backoff). If it fails: stop and DON'T save Excel.
    for color_key, acts in actions_by_color.items():
        has_add = any(a["op"] == "add" for a in acts)
        page = existing_by_color.get(color_key)

        if page is None and has_add:
            page = logic._with_backoff(
                notion.pages.create,
                parent={"database_id": nested_id},
                icon={"type": "emoji", "emoji": "⚙️"},
                properties={
                    "色": {"title": [{"text": {"content": color_key}}]},
                    "詳細": {"select": {"name": "ライン"}},
                    "品番": {"rich_text": []},
                    "数量": {"rich_text": []},
                    "c/t 秒": build_ct_prop(0, ct_is_number),
                },
            )
            existing_by_color[color_key] = page
            db_cache[color_key] = {"page": page, "parts": [], "qty_lines": [], "ct": 0}

        if page is None:
            continue

        props = page.get("properties", {}) or {}
        parts_rt = props.get("品番", {}).get("rich_text", []) or []
        qty_rt = props.get("数量", {}).get("rich_text", []) or []

        current_ct = int(db_cache.get(color_key, {}).get("ct") or read_ct_prop_from_page_props(props, ct_is_number) or 0)
        current_entries = parse_parts_lines(parts_rt)
        parts_map_local = build_parts_map()

        for act in acts:
            op = act["op"]
            trial = act["trial"]
            part = act["part"]
            qty = act["qty"]
            ct_int = max(0, int(act["ct"] or 0))
            d_str = (act["date"] or "").strip()

            full_text = part

            best_key = None
            best_len = -1
            for k in parts_map_local.keys():
                if full_text.startswith(k) and len(k) > best_len:
                    best_key = k
                    best_len = len(k)
            part_page_id = parts_map_local.get(best_key) if best_key else parts_map_local.get(part)

            if op == "add":
                already = any(((part_page_id and e["mention_id"] == part_page_id) or (not part_page_id and part and e["part_text"] and (part in e["part_text"]))) and (e["date"] == d_str) for e in current_entries)
                if already:
                    current_ct += ct_int
                    continue

                parts_rt = ensure_parts_has_separator(parts_rt)
                line_blocks: List[dict] = []

                if trial:
                    m = logic.TRIAL_CODE_RE.fullmatch(trial.strip())
                    if m:
                        line_blocks.append({"type": "text", "text": {"content": m.group(1)}, "annotations": {"bold": True, "color": "red_background"}})
                        line_blocks.append({"type": "text", "text": {"content": "・"}})
                    else:
                        line_blocks.append({"type": "text", "text": {"content": trial}})
                        line_blocks.append({"type": "text", "text": {"content": "・"}})

                logic._append_part_with_mention_and_modifier_auto(line_blocks, full_text, parts_map_local)

                if d_str:
                    md = format_date_mm_dd(d_str)
                    line_blocks.append({"type": "text", "text": {"content": f" {md}"}, "annotations": {"italic": True, "color": "green"}})

                parts_rt = trim_parts_trailing_newline(parts_rt + line_blocks)
                qty_rt = append_qty_green_italic(qty_rt, qty)
                current_ct += ct_int
                current_entries = parse_parts_lines(parts_rt)

            else:
                target_idx = find_part_line_index(parts_rt, part_page_id, full_text, d_str)
                if target_idx is None:
                    target_idx = find_part_line_index(parts_rt, part_page_id, full_text, "")
                if target_idx is not None:
                    parts_rt = remove_parts_line_at_index(parts_rt, target_idx)
                    qty_rt = remove_qty_at_index_if_green_italic_with_value(qty_rt, target_idx, qty)
                    current_ct = max(0, current_ct - ct_int)
                    current_entries = parse_parts_lines(parts_rt)

        if rich_text_is_effectively_empty(parts_rt) and rich_text_is_effectively_empty(qty_rt) and current_ct == 0:
            logic._with_backoff(notion.pages.update, page_id=page["id"], archived=True)
        else:
            logic._with_backoff(
                notion.pages.update,
                page_id=page["id"],
                properties={
                    "品番": {"rich_text": parts_rt},
                    "数量": {"rich_text": qty_rt},
                    "c/t 秒": build_ct_prop(int(current_ct), ct_is_number),
                },
            )

    # ✅ Now save Excel AFTER Notion succeeded
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(tmp)
    shutil.move(tmp, selected_file)

    wf_progress["products_loaded"] = True
    set_status("✅ Sync complete", THEMES[CURRENT_THEME]["status_ok"])
    log("Sync complete")
    set_progress(0)
