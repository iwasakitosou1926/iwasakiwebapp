"""
daily_workflow_generator.py
---------------------------

Daily Workflow Generator feature (Excel → Notion auto plan).

Responsibilities:
- Given a Notion calendar page and a target date, find the matching Excel file.
- Preprocess Excel (split N93・3F黒, etc.) using logic.py.
- Skip already-highlighted rows.
- Group remaining rows by color and send them to Notion『作業内容』DB.
- Highlight processed Excel rows.
- Update UI status / progress, and show dialogs.

Depends on:
- logic.py   for Notion + Excel helpers
- layout_themes.py  for colors
- Tkinter for messagebox + optional status/progress widgets
"""

from __future__ import annotations

import os
from typing import Optional, Dict, List
from datetime import date, datetime

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from layout_themes import THEMES, CURRENT_THEME
import logic

from logic import (
    preprocess_excel_split_n93_3f,
    read_excel_data_pandas,
    is_row_highlighted,
    COLOR_PAREN_RE,
    split_excel_color,
    build_parts_map,
    find_nested_databases,
    create_row_in_nested_db_auto,
    group_data_daily,
)

# Calendar helper (gap-safe): next available 4 days/pages
def get_calendar_pages_next4():
    return logic.get_calendar_pages_next4()

# =====================================================
# OPTIONAL UI WIDGET REFERENCES (wired by app_main)
# =====================================================


status_label: Optional[ttk.Label] = None
progress: Optional[ttk.Progressbar] = None

# Manual Excel override (set by app_main UI). If set and exists, Daily uses this file.
selected_excel_file: Optional[str] = None


def set_status(text: str, color: Optional[str] = None):
    """Update status label (if wired by app_main)."""
    global status_label
    if not status_label or not status_label.winfo_exists():
        return
    if color is None:
        color = THEMES[CURRENT_THEME]["text"]
    status_label.config(text=text, foreground=color)
    status_label.update_idletasks()


def set_progress(value: Optional[int] = None, mode: str = "determinate"):
    """Control progress bar (if wired by app_main)."""
    global progress
    if not progress or not progress.winfo_exists():
        return
    if mode == "indeterminate":
        progress.config(mode="indeterminate")
        progress.start(12)
    else:
        progress.stop()
        progress.config(mode="determinate")
        if value is not None:
            progress["value"] = value


def log(msg: str):
    """Delegate logging to logic.log (which app_main can hook into the UI console)."""
    logic.log(msg)


# =====================================================
# EXCEL FILE PICKER FOR DAILY GENERATOR
# =====================================================

def select_excel_for_daily(initial_dir: str, target_date: date) -> Optional[str]:
    """Prompt the user to select an Excel file for Daily Generator.

    This is used on Windows when folder auto-detection fails or when the user
    prefers manual selection (same pattern as Workflow Manager).
    """
    # Normalize / validate initial directory. Windows ignores invalid initialdir.
    try:
        getter = getattr(logic, "get_user_selected_excel_folder", None)
        if callable(getter):
            initial_dir = getter() or initial_dir
    except Exception:
        pass

    # UNC + slash normalization
    try:
        s = str(initial_dir).strip() if initial_dir else ""
    except Exception:
        s = ""

    candidates = []
    if s:
        candidates.append(s)
        candidates.append(s.replace("/", "\\"))
        if s.startswith("//"):
            candidates.append("\\\\" + s[2:].replace("/", "\\"))
        candidates = [os.path.normpath(x) for x in candidates]

    initial_dir_ok = None
    for c in candidates:
        try:
            if os.path.isdir(c):
                initial_dir_ok = c
                break
        except Exception:
            continue

    initial_dir = initial_dir_ok or (candidates[0] if candidates else None)
    if not initial_dir or not os.path.isdir(initial_dir):
        initial_dir = os.path.expanduser("~")

    try:
        log(f"[Daily] Excel picker initialdir = {initial_dir}")
    except Exception:
        pass

    try:
        title = f"Daily: Excelファイルを選択（{target_date.strftime('%Y-%m-%d')}）"
    except Exception:
        title = "Daily: Excelファイルを選択"

    fname = filedialog.askopenfilename(
        parent=tk._default_root,
        title=title,
        initialdir=initial_dir,
        filetypes=[
            ("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"),
            ("All files", "*.*"),
        ],
    )

    if not fname:
        return None

    # Persist folder so Workflow Manager + Daily share the same default
    try:
        picked_dir = os.path.dirname(fname)
        if picked_dir:
            setter = getattr(logic, "set_user_selected_excel_folder", None)
            if callable(setter):
                setter(picked_dir)
            try:
                logic.FOLDER_PATH = picked_dir
            except Exception:
                pass
    except Exception:
        pass

    # Remember manual selection for this session
    global selected_excel_file
    selected_excel_file = fname

    return fname


# =====================================================
# MAIN DAILY GENERATOR FUNCTION
# =====================================================

def run_daily_auto_for_page(selected_page: dict, target_date: date):
    """
    Daily Generator main function.

    Parameters
    ----------
    selected_page : dict
        Notion page object from the calendar DB (must contain 'id').
    target_date : date
        Date whose Excel file (MMDD_塗装*.xlsx) should be processed.

    Behavior
    --------
    - Finds Excel file for target_date under FOLDER_PATH.
    - Preprocesses mixed color rows (N93・3F黒 / 3分艶ブラック) via logic.preprocess_excel_split_n93_3f.
    - Reads data using pandas, skips already-highlighted rows.
    - Groups remaining data by color via logic.group_data_daily.
    - For each color group:
        * Creates a row in the nested『作業内容』DB inside selected_page.
        * Highlights corresponding Excel rows (first 10 columns) in yellow.
    - Saves Excel and updates status/progress.
    """
    set_status("Daily: Excel検索中…", THEMES[CURRENT_THEME]["text"])

    # 1) Choose Excel file (Windows-friendly)
    # If user selected an Excel file manually (from UI), prefer it.
    # If user selected an Excel file manually (from UI), prefer it.
    global selected_excel_file
    if selected_excel_file:
        if os.path.isfile(selected_excel_file):
            filepath = selected_excel_file
            log(f"[Daily] Using manually selected Excel file: {filepath}")
        else:
            log(f"[Daily] Manual Excel file not found: {selected_excel_file}")
            filepath = None
    else:
        filepath = None
        
    if not filepath or not os.path.isfile(filepath):
        messagebox.showwarning(
            "Daily",
            "Excelファイルが選択されていません。\n\nDaily Generatorを中止します。",
        )
        return

    log(f"[Daily] Using Excel file: {filepath}")

    # 2) Preprocess Excel to split rows like 'N93・3F黒'
    try:
        preprocess_excel_split_n93_3f(filepath)
    except Exception as e:
        log(f"[Daily] Preprocess error: {e}")
    
    # Add ロット# to duplicates so Notion lines are distinct
    try:
        logic.apply_lot_numbers_to_excel(filepath)
    except Exception as e:
        log(f"[LOT] Daily lot numbering failed: {e}")

    # 3) Load data with pandas + openpyxl (for highlighting)
    df = read_excel_data_pandas(filepath)
    wb = load_workbook(filepath)
    ws = wb.active

    headers = {c.value: i for i, c in enumerate(ws[1])}
    color_idx = headers.get("塗装色")
    full_idx = headers.get("子品番の正式名称")

    if color_idx is None:
        color_idx = 1  # fallback: 2nd column

    # 4) Detect already-highlighted rows so we can skip them in grouping
    highlighted_row_idxs = set()
    idx = -1  # pandas index is 0-based; ws.iter_rows(min_row=2) loops in same order
    for row in ws.iter_rows(min_row=2):
        idx += 1
        if is_row_highlighted(row):
            highlighted_row_idxs.add(idx)

    total_rows = len(df)
    keep_indices = [i for i in range(total_rows) if i not in highlighted_row_idxs]

    if not keep_indices:
        log("[Daily] All rows already highlighted — nothing to add.")
        df_to_group = df.iloc[0:0]
    else:
        if highlighted_row_idxs:
            log(f"[Daily] Skipping {len(highlighted_row_idxs)} already-highlighted rows.")
        df_to_group = df.iloc[keep_indices]

    # 5) Group rows by color / parts using existing logic
    groups = group_data_daily(df_to_group)
    parts_map = build_parts_map()

    # 6) Find nested『作業内容』DB inside the selected calendar page
    nested = find_nested_databases(selected_page["id"], "作業内容")
    if not nested:
        messagebox.showerror("Daily", "『作業内容』DBがページ内にありません。")
        return
    nested_id = nested[0]

    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 7) Create Notion rows per color
    step = 0
    total = max(1, len(groups))

    # --- BEFORE looping groups: detect duplicate part names across colors ---
    # Build a map: part_name → set of colors used
    part_color_usage = {}
    for c, info2 in groups.items():
        for p in info2["品番"]:
            part_color_usage.setdefault(p, set()).add(c)

    # Now inside the loop, use new_display_names logic
    for color, info in groups.items():
        step += 1
        set_progress(int(step / total * 100))
        log(f"[Daily] Adding {color} → Notion")

        # Build display names with color attached if part is used across multiple colors
        new_display_names = []
        for p in info["品番"]:
            if len(part_color_usage.get(p, [])) >= 2:
                # Attach color
                new_display_names.append(f"{p}({color})")
            else:
                new_display_names.append(p)

        try:
            create_row_in_nested_db_auto(
                nested_db_id=nested_id,
                color=color,
                display_names=new_display_names,
                finish_qty=info["数量"],
                total_cycle_time=info["作業時間(秒)"],
                parts_map=parts_map,
            )
        except Exception as e:
            log(f"[Daily] Notion write failed for color={color}: {type(e).__name__}: {e}")
            messagebox.showerror(
                "Daily",
                "Notionへの同期でエラーが発生しました。\n\n"
                f"色: {color}\n"
                f"エラー: {type(e).__name__}: {e}\n\n"
                "ネットワーク(社内Wi-Fi/VPN)やNotion接続を確認してください。",
            )
            set_progress(0)
            set_status("❌ Daily sync failed", THEMES[CURRENT_THEME]["status_warn"])
            return
        
        # 8) Highlight only the rows we just processed (i.e. NOT already highlighted)
        idx2 = -1
        for row in ws.iter_rows(min_row=2):
            idx2 += 1
            if idx2 in highlighted_row_idxs:
                continue  # skip pre-highlighted rows
            raw_color = ""
            if row[color_idx].value:
                raw_color = str(row[color_idx].value).strip()
            full_name = ""
            if full_idx is not None and row[full_idx].value:
                full_name = str(row[full_idx].value).strip()

            excel_colors = split_excel_color(raw_color, full_name)
            if color in excel_colors:
                for cell2 in row[:10]:
                    cell2.fill = yellow

    # 9) Save Excel (safe temp → replace)
    temp_path = filepath.replace(".xlsx", "_temp.xlsx")
    try:
        wb.save(temp_path)
        os.replace(temp_path, filepath)
    except Exception:
        wb.save(filepath)

    set_progress(0)
    set_status("✅ Daily Generator complete", THEMES[CURRENT_THEME]["status_ok"])
    messagebox.showinfo("Daily", "日々計画表作成が完了しました。")
    log("[Daily] 完了しました。")